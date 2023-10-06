# import libraries
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
import random
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from tabulate import tabulate
import itertools
import io
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
import sys

#generate user inputs
def generate_user_input(filename):
    workbook = Workbook()
    sheet = workbook.active
    sheet['A1'] = "Condition"
    sheet['B1'] = "A_data"
    sheet['C1'] = "B_data"

    for row in range(2, 52):  #specify the number of inputs + 2
        key = f"Condition{row - 1}"
        condition_A = random.choice(["yes", "no"])
        condition_B = random.choice(["yes", "no"])

        sheet[f'A{row}'] = key
        sheet[f'B{row}'] = condition_A
        sheet[f'C{row}'] = condition_B
    workbook.save(filename)



#read user Inputs
def read_user_input(filename):
    workbook = load_workbook(filename)
    sheet = workbook.active

    A_data = {}
    B_data = {}
    for row in range(2, 52):  #specify the number of inputs + 2
        key = sheet.cell(row=row, column=1).value
        condition_A = sheet.cell(row=row, column=2).value
        condition_B = sheet.cell(row=row, column=3).value
        A_data[key] = condition_A
        B_data[key] = condition_B

    return A_data, B_data


#count how many conditions are true for both A and B
def count_yes_conditions(A_data, B_data):
    A_yes_count = 0
    B_yes_count = 0
    for condition, value in A_data.items():
        if value == 'yes':
            A_yes_count += 1

    for condition, value in B_data.items():
        if value == 'yes':
            B_yes_count += 1

    return A_yes_count, B_yes_count




#generate ethical profiles

def generate_profiles(filename):
    workbook = Workbook()
    sheet = workbook.active

    sheet['A1'] = "Key"
    sheet['B1'] = "Description"
    sheet['C1'] = "A_profile"
    sheet['D1'] = "B_profile"

    for row in range(2, 52):  #specify the number of inputs + 2
        key = f"e{row - 1}"
        description = f"Principle{row - 1}"
        userA_value = random.randint(1, 5)
        userB_value = random.randint(1, 5)

        sheet[f'A{row}'] = key
        sheet[f'B{row}'] = description
        sheet[f'C{row}'] = userA_value
        sheet[f'D{row}'] = userB_value

    workbook.save(filename)





#read profiles
def read_profiles(filename):
    workbook = load_workbook(filename)
    sheet = workbook.active

    A_profiles = {}
    B_profiles = {}

    for row in range(2, 52):  #specify the number of inputs + 2
        key = sheet.cell(row=row, column=1).value
        description = sheet.cell(row=row, column=2).value
        A_profile = sheet.cell(row=row, column=3).value
        B_profile = sheet.cell(row=row, column=4).value

        A_profiles[key] = (description, A_profile)
        B_profiles[key] = (description, B_profile)

    return A_profiles, B_profiles





#generate tasks 

def generate_tasks(filename):
    workbook = Workbook()
    sheet = workbook.active
    sheet['A1'] = "Task"
    sheet['B1'] = "Description"

    for row in range(2, 52):  #specify the number of inputs + 2
        task = f"t{row - 1}"
        description = f"task{row - 1}"

        sheet[f'A{row}'] = task
        sheet[f'B{row}'] = description

    workbook.save(filename)



#assign tasks to a specified user

def assign_tasks(filename, username):
    workbook = load_workbook(filename)
    sheet = workbook.active
    user_dict = {}

    for row in range(2, 52):  #specify the number of inputs + 2
        task = sheet[f'A{row}'].value
        description = sheet[f'B{row}'].value
        user_dict[task] = description
    workbook.save(filename)

    return user_dict




#assign each task to minimum 1 or max 5 principles randomly

def create_task_implication(tasks, profile):
    task_implication = {}

    for task, task_description in tasks.items():
        num_principles = random.randint(1, 5) # specify the range of values
        principles = random.sample(list(profile.keys()), num_principles)
        task_implication[task] = principles

    return task_implication




# compute task ethical implication for each task
def calculate_task_value(profile, task_implication, user):
    result_df = pd.DataFrame(columns=task_implication.keys(), index=[user]) 

    for task, principles in task_implication.items():
        task_sum = sum(profile[principle][1] for principle in principles)

        result_df.loc[user, task] = task_sum

    return result_df



#each user input is asisgned to minimun 1 or maximum 5 principles to calculate possible number of offers
def assign_data_to_profiles(user_data, user_profiles):
    data_to_profiles = {item: [] for item in user_data}

    for item, value in user_data.items():
        num_profiles = random.randint(1, 5) #specify the range

        profiles = random.sample(list(user_profiles.keys()), num_profiles)
        data_to_profiles[item].extend(profiles)

    return data_to_profiles



#get all data into dictionaries
def get_all_data(A_data, B_data, A_profiles, B_profiles, A_task_implication, B_task_implication, result_df_A, result_df_B, A_data_to_profiles, B_data_to_profiles, A_tasks, B_tasks):

    data= {
    'A' : A_data,
    'B' : B_data
    }
    profiles= {
        'A' : A_profiles,
        'B' : B_profiles
    }
    Task_implications = {
        'A' : A_task_implication,
        'B' : B_task_implication
        }
    Task_value = {
        'A' : result_df_A,
        'B' : result_df_B
        }
    data_associated_to_principles = {
        'A' : A_data_to_profiles,
        'B' : B_data_to_profiles
        }
    Total_tasks = {
        'A' : A_tasks,
        'B' : B_tasks
        }
    return data, profiles, Task_implications, Task_value, data_associated_to_principles, Total_tasks



#assign data to agents A and B
def assign_data_to_agent(data, profiles, Task_implications, Task_value, data_associated_to_principles, Total_tasks):
    agents = {}
    for agent in data.keys():
        agents[agent] = {}
        agents[agent]['data'] = data[agent]
        agents[agent]['profile'] = profiles[agent]
        agents[agent]['Task_implications'] = Task_implications[agent]
        agents[agent]['Task_value'] = Task_value[agent]
        agents[agent]['data_associated_to_principles'] = data_associated_to_principles[agent]
        agents[agent]['Total_tasks'] = Total_tasks[agent]
       
    return agents



#to randomly select sender and receiver
def select_sender_receiver(agents):
    sender = random.choice(list(agents.keys()))
    receiver = random.choice(list(agents.keys()))
    while sender == receiver:
        receiver = random.choice(list(agents.keys()))
    return sender, receiver



#offer generation function
def generate_offers(agent, agents):
    agent_all_offers = {}

    agent_data = agents[agent]['data']
    agent_data_associated_to_principles = agents[agent]['data_associated_to_principles']
    agent_task_implications = agents[agent]['Task_implications']
    agent_task_value = agents[agent]['Task_value']

    # Loop through the data and check conditions with 'yes' value
    for condition, value in agent_data.items():
        if value == 'yes':
            tasks_for_condition = []

            # Find the principles associated with this condition
            associated_principles = agent_data_associated_to_principles[condition]

            # Find tasks based on principles from Task Implications
            for task, implications in agent_task_implications.items():
                if any(implication in associated_principles for implication in implications):
                    tasks_for_condition.append(task)

            offer_utility = int(sum(float(agent_task_value[task].iloc[0]) for task in tasks_for_condition))
            agent_all_offers[condition] = {
                'tasks': tasks_for_condition,
                'utility': offer_utility
            }

    return agent_all_offers



#SWITCH AGENT ROLE
def switch_role(sender, receiver):
    temp = sender
    sender = receiver
    receiver = temp
    return sender, receiver


#SWITCH DATA
def switch_data(S_offer, S_utility, S_all_offers, S_remaining_offers, R_offer, R_utility, R_all_offers, R_remaining_offers):
    dummy_offer = R_offer
    dummy_utiltiy = R_utility
    dummy_all_offers = R_all_offers
    dummy_remaining_offers = R_remaining_offers

    R_offer = S_offer
    R_utility = S_utility
    R_all_offers = S_all_offers
    R_remaining_offers = S_remaining_offers

    S_offer = dummy_offer
    S_utility = dummy_utiltiy
    S_all_offers = dummy_all_offers
    S_remaining_offers = dummy_remaining_offers

    return S_offer, S_utility, S_all_offers, S_remaining_offers, R_offer, R_utility, R_all_offers, R_remaining_offers


#choose from first 10 percent of offers with max utility
def find_best_offer(offers):
    sorted_offers = sorted(offers.items(), key=lambda x: x[1]['utility'], reverse=True)

    # Calculate the number of offers to consider (10% of total)
    num_offers_to_consider = len(sorted_offers) // 10
    selected_offer = random.choice(sorted_offers[:num_offers_to_consider])

    offer = selected_offer[1]['tasks']
    utility = selected_offer[1]['utility']

    remaining_offers = {condition: off for condition, off in offers.items() if condition != selected_offer[0]}
    remaining_sorted_offers = sorted(remaining_offers.items(), key=lambda x: x[1]['utility'], reverse=True)

    return offer, utility, remaining_sorted_offers, sorted_offers


#dummy offer generation function to use in round 2 when agents switch data
def updated_dummy_best_offer(remaining_offers):
    # Calculate the number of offers to consider (10% of the length)
    num_offers_to_consider = len(remaining_offers) // 10

    if num_offers_to_consider < 1:
        # If there are fewer than 10% offers, choose one randomly
        selected_offer = random.choice(remaining_offers)
    else:
        selected_offers = remaining_offers[:num_offers_to_consider]
        selected_offer = random.choice(selected_offers)

    offer = selected_offer[1]['tasks']
    utility = selected_offer[1]['utility']
    remaining_offers.remove(selected_offer)

    return offer, utility, remaining_offers


#utility function
def calculate_new_random_utility(Sender_offer, receiver, agents, Receiver_offer):
    sender_utility = {}
    receiver_utility = {}
    Sender_task_value_sum = 0
    R_task_value_sum = 0

    for tasks in Sender_offer:
        Sender_task_value_sum += agents[receiver]['Task_value'].get(tasks, 0) 
    for tasks in Receiver_offer:
        R_task_value_sum += agents[receiver]['Task_value'].get(tasks, 0) 

    utility = Sender_task_value_sum - R_task_value_sum
    sender_utility = int(Sender_task_value_sum)
    receiver_utility = int(R_task_value_sum)
    utility = int(utility)

    return sender_utility, receiver_utility, utility

################ to change tactic

#find max utility offer for receiver
def find_max_utility_offer(offers):
    # Sort offers by utility in descending order
    R_sorted_offers = sorted(offers.items(), key=lambda x: x[1]['utility'], reverse=True)

    # Select the first (highest value) offer
    selected_offer = R_sorted_offers[0]

    # Extract tasks and utility from the selected offer
    R_offer = selected_offer[1]['tasks']
    R_utility = selected_offer[1]['utility']

    # Create a dictionary for the remaining offers
    remaining_offers = {condition: offer for condition, offer in offers.items() if condition != selected_offer[0]}

    # Sort the remaining offers and convert them to a list
    R_remaining_offers = sorted(remaining_offers.items(), key=lambda x: x[1]['utility'], reverse=True)

    return R_offer, R_utility, R_remaining_offers, R_sorted_offers



#find max utility offer for sender
def find_max_utility_offer(offers):
    
    S_sorted_offers = sorted(offers.items(), key=lambda x: x[1]['utility'], reverse=True)
    selected_offer = S_sorted_offers[0]

    S_offer = selected_offer[1]['tasks']
    S_utility = selected_offer[1]['utility']

    remaining_offers = {condition: offer for condition, offer in offers.items() if condition != selected_offer[0]}
    S_remaining_offers = sorted(remaining_offers.items(), key=lambda x: x[1]['utility'], reverse=True)

    return S_offer, S_utility, S_remaining_offers, S_sorted_offers


# dummy function to use in round 2 and more
def dummy_max_offer(remaining_offers):

    # Find the offer with the highest utility
    max_utility_offer = max(remaining_offers, key=lambda x: x[1]['utility'])
    
    # Extract tasks and utility from the selected offer
    R_offer = max_utility_offer[1]['tasks']
    R_utility = max_utility_offer[1]['utility']

    # Remove the selected offer from the remaining offers
    remaining_offers.remove(max_utility_offer)

    return R_offer, R_utility, remaining_offers

#find random offer for receiver
def find_random_offer(offers):

    R_all_offers = offers
    selected_offer = random.choice(list(offers.items()))

    R_offer = selected_offer[1]['tasks']
    R_utility = selected_offer[1]['utility']
    R_remaining_offers = {condition: offer for condition, offer in offers.items() if condition != selected_offer[0]}

    return R_offer, R_utility, R_remaining_offers, R_all_offers



#find random offer for sender
def find_random_offer(offers):

    S_all_offers = offers
    selected_offer = random.choice(list(offers.items()))

    S_offer = selected_offer[1]['tasks']
    S_utility = selected_offer[1]['utility']
    S_remaining_offers = {condition: offer for condition, offer in offers.items() if condition != selected_offer[0]}

    return S_offer, S_utility, S_remaining_offers, S_all_offers



##### dummy random offer to use in round 2 and more
def dummy_random_offer(remaining_offers):

    selected_condition = random.choice(list(remaining_offers.keys()))
    R_offer = remaining_offers[selected_condition]['tasks']
    R_utility = remaining_offers[selected_condition]['utility']
    del remaining_offers[selected_condition]

    return R_offer, R_utility, remaining_offers

